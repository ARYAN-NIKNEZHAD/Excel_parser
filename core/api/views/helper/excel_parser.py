import openpyxl
from core.models import Device, DeviceType, Value, Series, ExcelFile


def parse_excel_file(file):
    """
    Parses the provided Excel file and updates the database with extracted data.
    
    Args:
    - file (str): The path to the Excel file to parse.
    
    Returns:
    - ExcelFile: The ExcelFile object created for the parsed file.
    """
    # Load the Excel workbook in read-only mode
    wb = openpyxl.load_workbook(file, read_only=True)

    # Create a new ExcelFile object for the parsed file
    excel = ExcelFile.objects.create(file=file)

    # Parse the 'Devices' sheet and create Device objects
    device_objects = {}
    for row in wb['Devices'].iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            device_name = row[0].capitalize()
            device = Device.objects.create(excel_file=excel, name=device_name)
            device_objects[device_name] = device

    # Create DeviceType objects
    device_types_to_create = []
    for device_name, device in device_objects.items():
        for row in wb[f'{device_name}s'].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                device_type_name = row[0]
                device_types_to_create.append(DeviceType(excel_file=excel, device=device, name=device_type_name))
    DeviceType.objects.bulk_create(device_types_to_create)


    # Create Series and Value objects
    series_to_create = []
    values_to_create = []
    for device_name, device in device_objects.items():
        device_type_mapping = {dt.name: dt for dt in DeviceType.objects.filter(device=device)}
        for row in wb[f'{device_name}-Series'].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                device_type_name, series_name, *values = row
                device_type = device_type_mapping.get(device_type_name)
                if device_type:
                    series_to_create.append(Series(excel_file=excel, device_type=device_type, name=series_name))

                    # Prepare Value objects for the series
                    series_index = len(series_to_create) - 1
                    for value in values:
                        if value is not None:
                            values_to_create.append(Value(excel_file=excel, series=series_to_create[series_index], value=value))

    # Bulk create Series and Value objects
    Series.objects.bulk_create(series_to_create)
    Value.objects.bulk_create(values_to_create)

    return excel