import openpyxl
from rest_framework.serializers import ValidationError

def update_excel_series_value(excel_file, series_name, device_type, value):
    """
    Updates the specified series and its associated data in both the database and the Excel file.

    Args:
    - excel_file (ExcelFile): The existing ExcelFile object.
    - series_name (str): The name of the series to update.
    - device_type (DeviceType): The device type associated with the series.
    - value (str): The new value to be added or changed.
    
    Raises:
    - ValidationError: If there is no empty cell in the series to add the value.

    """
    # Load the workbook in write mode
    wb = openpyxl.load_workbook(excel_file.file.path, read_only=False)

    # Access the specific worksheet corresponding to the device type
    sheet = wb[f"{device_type.device.name}-Series"]

    count = 1  # Initialize a counter to track the row number
    for row in sheet.iter_rows(min_row=2, values_only=True):
        count += 1  # Increment the row counter
        # Check if the row corresponds to the specified device type and series name
        if row and row[0] == device_type.name and row[1] == series_name:
            # Find the last non-empty cell in the row
            idx = len(row)
            # Raise ValidationError if there is no empty cell to add the value
            if row[idx - 1] is not None:
                raise ValidationError(
                    {"Error": "There is no empty cell in the series to add value"}
                )
            while idx > 0 and row[idx - 1] is None:
                idx -= 1

            # Get the index of the last non-empty cell
            row_index = row.index(row[idx])
            # Update the value in the next empty cell in the row
            sheet.cell(row=count, column=row_index + 1, value=value)
            break

    # Save the changes to the Excel file
    wb.save(excel_file.file.path)