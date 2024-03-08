import openpyxl

def update_excel_series_value_change(excel_file, series_name, device_type, value):
    """
    Updates the specified series and its associated data in both the database and the Excel file.
    
    Args:
    - excel_file (ExcelFile): The existing ExcelFile object.
    - series_name (str): The name of the series to update.
    - device_type (DeviceType): The device type associated with the series.
    - value (str): The new value to be added or changed.
    """
    # Load the workbook in write mode
    wb = openpyxl.load_workbook(excel_file.file.path, read_only=False)

    # Access the specific worksheet corresponding to the device type
    sheet = wb[f"{device_type.device.name}-Series"]

    # Initialize row count
    count = 1

    # Iterate through each row in the worksheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        count += 1
        # Check if the row corresponds to the specified device type and series name
        if row and row[0] == device_type.name and row[1] == series_name:
            # Find the column index for the specified series
            idx = row.index(series_name)

            # Access the cell corresponding to the specified series and row
            cell = sheet.cell(row=count, column=idx + 1)

            # Update the value of the cell with the new value
            cell.value = value
            break

    # Save the changes to the Excel file
    wb.save(excel_file.file.path)